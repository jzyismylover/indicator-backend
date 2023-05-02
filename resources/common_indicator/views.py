import hashlib
import uuid
from openpyxl import Workbook
from io import BytesIO
from flask import send_file
from flask_restful import Resource, reqparse, abort, inputs, request
from config.redis import get_dyn_data, mark_dyn_data, delete_dyn_data
from resources.common_indicator.util import *
from utils.jwt import check_premission
from utils.json_response import make_error_response, make_success_response

parser = reqparse.RequestParser()
# actually it must be application/json
parser.add_argument('list', location='form')
parser.add_argument('lg_type', type=str, location='form')
parser.add_argument('lg_text', type=str, location='form')
parser.add_argument('requireSplit', type=inputs.boolean, location='form')


# 文本hash散列函数
def generateHash(text: str):
    hash_model = hashlib.md5()
    hash_model.update(text.encode('utf-8'))
    return hash_model.hexdigest()


def getParams(parser=parser, id=None, use_redis=True) -> CommonIndicatorHandler:
    def getParamsHandler(lg_text, lg_type, id, requireSplit=False):
        hash_value = generateHash(lg_type + lg_text)
        # replace some \n | \r\b | \\n chars
        lg_text = lg_text.strip().replace('\n', '').replace('\r', '').replace('\\n', '')
        handler = getLanguageHandler(
            lg_text,
            lg_type,
            id,
            hash_value=hash_value,
            use_redis=use_redis,
            requireSplit=requireSplit,
        )
        return handler

    # params = request.get_json()
    params = parser.parse_args()
    if params['requireSplit'] is None:
        requireSplit = False
    else:
        requireSplit = params['requireSplit']
    # 兼容单文本传入
    if params['list'] is None:
        params = parser.parse_args()
        lg_type = params['lg_type']
        lg_text = params['lg_text']
        # invalid empty content
        if lg_text is '':
            return abort(400, **make_error_response(msg='内容为空'))
        handler = getParamsHandler(lg_text, lg_type, id, requireSplit)
        return handler
    # 多文本传入
    else:
        lists = params['list']
        handlers = []
        for item in lists:
            handler = getParamsHandler(
                item['lg_text'], item['lg_type'], id, requireSplit
            )
            handlers.append(handler)
        return handlers


def getLanguageHandler(
    lg_text, lg_type, id, *, hash_value='', use_redis, requireSplit
):
    if use_redis is False:
        return CommonIndicatorHandler(
            text=lg_text, lg_type=lg_type, requireSplit=requireSplit
        )
    try:
        # connect error redis
        handler = get_dyn_data(hash_value)
        if handler == None:
            handler = CommonIndicatorHandler(
                text=lg_text, lg_type=lg_type, requireSplit=requireSplit
            )
            mark_dyn_data(hash_value, handler)
    except Exception as e:
        handler = CommonIndicatorHandler(
            text=lg_text, lg_type=lg_type, requireSplit=requireSplit
        )

    return handler


def handleIndicatorReturn(**kargs):
    return {'data': {kargs['type']: kargs['value']}}


# 多语种分词接口
class Tokenizen(Resource):
    def post(self):
        handler = getParams(use_redis=False)
        return {"data": {"ans": handler.handleTokenizen()}}


# 多语种词性标注接口
class SpeechTagging(Resource):
    def post(self):
        handler = getParams(use_redis=False)
        return {"data": {"ans": handler.handleSpeechTagging()}}


"""
具体指标计算视图
"""


def handleIndicatorCalculate(handler, fn):
    if isinstance(handler, list) is False:
        ans = fn(handler)
    else:
        ans = []
        for _ in handler:
            _ans = fn(_)
            ans.append(_ans)

    return ans


class TotalWords(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        total_words = handleIndicatorCalculate(
            handler, lambda handler: len(handler.words)
        )

        return handleIndicatorReturn(value=total_words, type='Words(总词数)')


class DictWords(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        dict_words = handleIndicatorCalculate(
            handler, lambda handler: len(handler.frequency)
        )

        return handleIndicatorReturn(value=dict_words, type='Dicts(词典数)')


class HapaxWords(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        dict_words = handleIndicatorCalculate(
            handler, lambda handler: len(handler.hapax)
        )

        return handleIndicatorReturn(value=dict_words, type='hapaxWords(单现词数)')


class TTRValue(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        ans = handleIndicatorCalculate(handler, lambda handler: handler.getTTRValue())

        return handleIndicatorReturn(value=ans, type='TTR(型例比)')


class HPoint(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        h_point = handler.getHPoint()

        return handleIndicatorReturn(value=h_point, type='Hpoint(h点)')


class EntropyValue(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])

        H = handler.getEntroyValue()

        return handleIndicatorReturn(value=H, type='Entropy(文本熵)')


"""
暂时省略 Average Tokens Length & Token Length Frequency Specturm
"""


class R1Value(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        R1 = handler.getR1Value()

        return handleIndicatorReturn(value=R1, type='R1(词汇丰富度)')


class RRValue(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        RR = handler.getRRValue()

        return handleIndicatorReturn(value=RR, type='RR(重复率)')


class RRmcValue(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        RRmc = handler.getRRmcValue()

        return handleIndicatorReturn(value=RRmc, type='RRmc(相对重复率)')


class TCValue(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        Tr = handler.getTCValue()

        return handleIndicatorReturn(value=Tr, type='TC(主题集中度)')


class SecondaryTCValue(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        Tr = handler.getSecondTCValue()

        return handleIndicatorReturn(value=Tr, type='Secondary(次级主题集中度)')


class ActivityValue(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        activity = handler.getAcitvityValue()

        return handleIndicatorReturn(value=activity, type='Activity(活动度)')


class DescriptivityValue(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        descriptivity = handler.getDescriptivityValue()

        return handleIndicatorReturn(value=descriptivity, type='Descriptivity(描写度)')


class LValue(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        L = handler.getLValue()

        return handleIndicatorReturn(value=L, type='L(文本中词的秩序分布欧氏距离)')


class CurveLengthValue(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        R = handler.getCurveLengthValue()

        return handleIndicatorReturn(value=R, type='Curve Length R Index(文本中词的秩序分布R指数)')


class LambdaValue(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        lambda_v = handler.getLambdaValue()

        return handleIndicatorReturn(value=lambda_v, type='lambda(文本Lambda值)')


class AdjustedModuleValue(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        A = handler.getAdjustModuleValue()

        return handleIndicatorReturn(value=A, type='Adjusted Modules(校正模数)')


class GiniValue(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        G = handler.getGiniValue()

        return handleIndicatorReturn(value=G, type='G(基尼系数)')


class R4Value(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        R4 = handler.getR4Value()

        return handleIndicatorReturn(value=R4, type='R4(基尼系数补数)')


class HapaxValue(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        rate = handler.getHapaxValue()

        return handleIndicatorReturn(value=rate, type='Hapax Percentage(单现词占比)')


class WriterView(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        cosa = handler.getWriterView()

        return handleIndicatorReturn(value=cosa, type='Writer\'s View(作者视野)')


class VerbDistance(Resource):
    @check_premission
    def post(self, info):
        handler = getParams(parser=parser, id=info['user_id'])
        verb_V = handler.getVerbDistance()

        return handleIndicatorReturn(value=verb_V, type='Verb Distances(动词间距)')


# class ZipfTest(Resource):
#     @check_premission
#     def post(self, info):
#         handler = getParams(parser=parser, id=info['user_id'])
#         zipf = handler.getZipf()
#         current_fields = {str(x): fields.Float for x in zipf.keys()}

#         return handleIndicatorReturn(
#             value=zipf, fields=current_fields, type='Zipf Test'
#         )


class ALLCommonIndicator(Resource):
    @check_premission
    def post(self, info):
        import datetime

        start_time = datetime.datetime.now()
        handler = getParams(parser=parser, id=info['user_id'])
        h = handler.getHPoint()
        Activity = handler.getAcitvityValue()
        G = handler.getGiniValue()

        end_time = datetime.datetime.now()
        print(f'parser time ${end_time - start_time}')
        

        data = {
            'Words(总词数)': len(handler.words),
            'Dicts(词典数)': len(handler.frequency),
            # 'HapaxWords(单现词数)': len(handler.hapax),
            'Hapax Percentage(单现词比例)': len(handler.hapax) / len(handler.words),
            'TTR(型例比)': handler.getTTRValue(),
            'HPoint(h点)': h,
            'Entropy(文本熵)': handler.getEntroyValue(),
            'R1(词汇丰富度)': handler.getR1Value(),
            'RR(重复率)': handler.getRRValue(),
            'RRmc(相对重复率)': handler.getRRmcValue(),
            'TC(主题集中度)': handler.getTCValue(),
            'SecondTc(次级主题集中度)': handler.getSecondTCValue(),
            'Activity(活动度)': Activity,
            'Descriptivity(描写度)': 1 - Activity,
            'L(文本中词的秩序分布欧氏距离)': handler.getLValue(),
            'Curver Length R Index(文本中词的秩序分布R指数)': handler.getCurveLengthValue(),
            'Lambda(文本Lambda值)': handler.getLambdaValue(),
            'G(基尼系数)': G,
            'R4(基尼系数补数)': 1 - G,
            'Writer\'s View(作者视野)': handler.getWriterView(),
            'Verb Distances(动词间距)': handler.getVerbDistance(),
        }
        hash_id = uuid.uuid4().__repr__()[6:-3]
        mark_dyn_data(hash_id, data)
        data['hash_value'] = hash_id

        return make_success_response(data=data)


class DownloadIndicatorsIntoExcel(Resource):
    @check_premission
    def post(self, info):
        if 'hash_value' not in request.form:
            return make_error_response(msg='hash_value is not is required'), 400
        hash_value = request.form['hash_value']
        data = get_dyn_data(hash_value)
        # delete_dyn_data(hash_value) ## 存在多次下载的情况不能删除

        wb = Workbook()
        ws = wb.active
        ws['A1'] = '指标名'
        ws['B1'] = '指标值'

        i = 2
        for row in data.keys():
            if row == 'hash_value':
                continue
            ws.cell(row=i, column=1).value = row
            ws.cell(row=i, column=2).value = data[row]
            i += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # with NamedTemporaryFile() as tmp:
        #     wb.save('jzyismylover.xlsx')
        #     tmp.seek(0)
        #     stream = tmp.read()
        print("{} b".format(len(output.getvalue())))

        fv = send_file(
            output,
            download_name='indicator.xlsx',
            as_attachment=True,
            conditional=True,
        )
        # fv.headers[
        #     'Content-Type'
        # ] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        # fv.headers["Cache-Control"] = "no-cache"
        # fv.headers['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(
        #     'jzyismylover'
        # )

        return fv

class DeleteIndicatorExcelHashCache(Resource):
    @check_premission
    def delete(self, info):
        if 'hash_value' not in request.form:
            return make_error_response(msg='hash_value param is required'), 400
        hash_value = request.form['hasg_value']
        delete_dyn_data(hash_value)