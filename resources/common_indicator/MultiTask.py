import json
import uuid
from io import BytesIO
from openpyxl import Workbook
from sqlalchemy import select
from flask_restful import Resource, request
from resources.Language import LanguageRec
from werkzeug.datastructures import FileStorage
from utils.jwt import check_premission
from utils.json_response import make_success_response
from resources.common_indicator.util import CommonIndicatorHandler
from config._celery import celery
from flask_mail import Message
from config.mail import mail
from resources.constant import (
    COMMON_INDICATOR_HANDLER_MAPPING,
)
from config.db import engine
from config.db.user import User


language_ins = LanguageRec()


@celery.task(name='resources.common_indicator.MultiTask.parse_files')
def parse_files(files, email):
    files = json.loads(files)
    lis = []
    for file in files:
        file = json.loads(file)
        filename = file['filename']
        contentType = file['content_type']
        # data = file['data'].encode('uft-8')
        data = bytes.fromhex(file['data'])
        file_storage = FileStorage(BytesIO(data), filename, content_type=contentType)
        lg_type, lg_content = language_ins.parse_file(file_storage)

        model = CommonIndicatorHandler(lg_content, lg_type)
        ans = dict()
        ans['filename'] = filename

        indicators = COMMON_INDICATOR_HANDLER_MAPPING.keys()
        for _ in indicators:
            if _ not in COMMON_INDICATOR_HANDLER_MAPPING:
                continue
            func = getattr(model, COMMON_INDICATOR_HANDLER_MAPPING[_], None)
            # 非增量指标累计
            score = func()
            ans[_] = score

        lis.append(ans)

    try:
        wb = Workbook()
        for index, _ in enumerate(lis):
            ws = wb.create_sheet(_['filename'], index=index)
            ws['A1'] = '指标名'
            ws['B1'] = '指标值'
            i = 2
            for row in indicators:
                if row == 'hash_value':
                    continue
                ws.cell(row=i, column=1).value = row
                ws.cell(row=i, column=2).value = _[row]
                i += 1
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        message = Message(
            subject='指标提取结果', body=f'请在附件查收指标提取结果', recipients=[email]
        )
        filename = f'{uuid.uuid4().__repr__()[6:-3]}.xlsx'
        message.attach(filename, 'application/octet-stream', output.read())
        mail.send(message)
        return json.dumps(
            make_success_response(msg='附件发送成功')
        )
    except Exception as e:
        return {'error': str(e)}


class MultiProcessCommonIndicator(Resource):
    def __init__(self) -> None:
        self.files = request.files.getlist('files')

    @check_premission
    def post(self, info):
        # search for user email
        with engine.connect() as conn:
            stmt = (
                select(User).where(User.id == info['user_id'])
            )
            rows = conn.execute(stmt)
            row = list(rows)[0]
            email = row.email

        files = []
        for file in self.files:
            data = {
                'filename': file.filename,
                'content_type': file.content_type,
                'content_length': file.content_length,
                'data': file.read().hex()
                # 'data': file.read()
            }
            files.append(json.dumps(data))
        files = json.dumps(files)
        task = parse_files.delay(files, email)
        return make_success_response(msg='success', data={'id': task.id})


class GetStatue(Resource):
    def get(self):
        id = request.form['id']
        result = parse_files.AsyncResult(id)
        return make_success_response(data=json.loads(result.info))

class StoreMultiFiles(Resource):
    def post(self):
        file = request.files.get