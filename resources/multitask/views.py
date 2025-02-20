import json
import uuid
import time
from datetime import datetime
from io import BytesIO
from flask import send_file
from flask_mail import Message
from openpyxl import Workbook
from sqlalchemy import select, insert, delete
from flask_restful import Resource, request
from resources.Language import LanguageRec
from werkzeug.datastructures import FileStorage
from resources.common_indicator.util import CommonIndicatorHandler
from resources.constant import (
    COMMON_INDICATOR_HANDLER_MAPPING,
)
from config._schedule import scheduler
from config._celery import celery
from config.mail import mail
from config.db import engine
from config.db.user import User
from config.db.tasks import Task
from utils.jwt import check_premission
from utils.json_response import make_success_response

language_ins = LanguageRec()


def generateBinaryExcelData(lis, indicators):
    wb = Workbook()
    for index, _ in enumerate(lis):
        ws = wb.create_sheet(_['filename'], index=index)
        ws['A1'] = '指标名'
        ws['B1'] = '指标值'
        i = 2
        for row in indicators:
            if row == 'hash_value':
                continue
            ws.cell(row=i, column=1).value = row
            ws.cell(row=i, column=2).value = _[row]
            i += 1
    return wb


def generateBinaryRawTextData(texts: list):
    content = ' '.join(texts)
    output = BytesIO()
    output.write(content.encode())
    output.seek(0)

    return output


# 设置定时任务，定时清除过期redis缓存处理数据
@scheduler.task('cron', id='clear_redis', hour='*/2')
def clear_redis_cache():
    _now = time.gmtime()
    _now = f'{_now.tm_year}-{_now.tm_mon}-{_now.tm_mday}'
    _now = datetime.strptime(_now, '%Y-%m-%d')
    with engine.connect() as conn:
        stmt = (
            select(Task)
        )
        taskLists = list(conn.execute(stmt))
    with engine.begin() as conn:
        for _ in taskLists:
              _pre = datetime.strptime(_.update_time, '%Y-%m-%d')
              _time = _now - _pre
              if _time.days >= 1:
                stmt = delete(Task).where(Task.id == _.id)
                conn.execute(stmt)

# 设置celery后台任务，异步后台执行密集计算任务
@celery.task(name='parse_files')
def parse_files(files, email):
    files = json.loads(files)
    lis = []
    for file in files:
        file = json.loads(file)
        filename = file['filename']
        contentType = file['content_type']
        # data = file['data'].encode('uft-8')
        data = bytes.fromhex(file['data'])
        file_storage = FileStorage(BytesIO(data), filename, content_type=contentType)
        lg_type, lg_content = language_ins.parse_file(file_storage)

        model = CommonIndicatorHandler(lg_content, lg_type)
        ans = dict()
        ans['filename'] = filename
        ans['content'] = ' '.join(model.words)
        
        # 计算指定指标
        indicators = COMMON_INDICATOR_HANDLER_MAPPING.keys()
        for _ in indicators:
            if _ not in COMMON_INDICATOR_HANDLER_MAPPING:
                continue
            func = getattr(model, COMMON_INDICATOR_HANDLER_MAPPING[_], None)
            score = func()
            ans[_] = score

        lis.append(ans)

    try:
        # 发送结果附件到邮箱
        wb = generateBinaryExcelData(lis, indicators)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        message = Message(subject='指标提取结果', body=f'请在附件查收指标提取结果', recipients=[email])
        filename = f'{uuid.uuid4().__repr__()[6:-3]}.xlsx'
        message.attach(filename, 'application/octet-stream', output.read())
        mail.send(message)
        return json.dumps(make_success_response(data=lis))
    except Exception as e:
        return {'error': str(e)}


class MultiProcessCommonIndicator(Resource):
    def __init__(self) -> None:
        self.files = request.files.getlist('files')

    @check_premission
    def post(self, info):
        # 搜索用户邮件
        with engine.connect() as conn:
            stmt = select(User).where(User.id == info['user_id'])
            rows = conn.execute(stmt)
            row = list(rows)[0]
            email = row.email

        files = []
        for file in self.files:
            data = {
                'filename': file.filename,
                'content_type': file.content_type,
                'content_length': file.content_length,
                # 先将二进制表示为十六进制码
                'data': file.read().hex()
            }
            files.append(json.dumps(data))
        files = json.dumps(files)
        task = parse_files.delay(files, email)

        # 新增任务列表
        _time = time.gmtime()
        update_time = f'{_time.tm_year}-{_time.tm_mon}-{_time.tm_mday}'
        with engine.connect() as conn:
            conn.execute(
                insert(Task),
                [
                    {
                        'id': task.id,
                        'user_id': info['user_id'],
                        'update_time': update_time,
                    }
                ],
            )
            conn.commit()
        return make_success_response(msg='success', data={'id': task.id})


class GetTaskInfo(Resource):
    def get(self):
        id = request.args['id']
        result = parse_files.AsyncResult(id)
        if result.status != 'SUCCESS':
            info = json.dumps({'status': result.status})
        else:
            info = result.info

        return json.loads(info)


class GetTaskLists(Resource):
    @check_premission
    def get(self, info):
        user_id = info['user_id']
        with engine.connect() as conn:
            results = conn.execute(select(Task).where(Task.user_id == user_id))
            results = list(results)
        
        # celery 任务默认有效时间为 24 hour(可在 config/_celery自定义)
        ans = []
        for _ in results:
            info = parse_files.AsyncResult(_.id)
            ans.append(
                {'id': _.id, 'status': info.status, 'update_time': _.update_time}
            )

        return make_success_response(data=ans)


class DownloadTask(Resource):
    @check_premission
    def post(self, info):
        id = request.form['id']
        info = parse_files.AsyncResult(id).info
        info = json.loads(info)
        wb = generateBinaryExcelData(
            info['data'], COMMON_INDICATOR_HANDLER_MAPPING.keys()
        )
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        fv = send_file(
            output,
            download_name='indicator.xlsx',
            as_attachment=True,
            conditional=True,
        )
        return fv


# 导出分词处理结果词典
class DownloadTaskDirectory(Resource):
    @check_premission
    def post(self, info):
        id = request.form['id']
        info = parse_files.AsyncResult(id).info
        info = json.loads(info)['data']
        _ = map(lambda row: row['content'], info)
        output = generateBinaryRawTextData(list(_))
        fv = send_file(
            output, download_name='dict.txt', as_attachment=True, conditional=True
        )

        return fv